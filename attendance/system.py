from openpyxl import Workbook, load_workbook
from django.shortcuts import get_object_or_404
from .models import Account, CheckSheet
from django.utils import timezone
from mysite2.settings import BASE_DIR
from datetime import datetime, timedelta, date
import calendar
import os

wageTime_dir = BASE_DIR/'excel_sheets/wageTime_sheets 2023'
dailyReport_sheet = BASE_DIR/'excel_sheets/QB Daily Report.xlsx'
salesReport_sheet = 'C:/Users/anko1/Documents/web_django/Sales Report QB 2023.xlsx'
staff_sheet = BASE_DIR/'excel_sheets/staff.xlsx'

def AttendanceTimeCalc(time):
    if(time.minute >= 1 and time.minute <= 30):
        time = time.replace(minute=30)
    elif(time.minute >= 31 and time.minute <= 59):
        time = time.replace(hour=time.hour+1, minute=00)
    
    return time

def LeavingTimeCalc(time):
    if(time.minute >= 16 and time.minute <= 45):
        time = time.replace(minute=30)
    else:
        time = time.replace(minute=0)
        if(time.minute >= 46):
            time = time.replace(hour=time.hour+1)

    return time


def WriteLeaving(account):

    now = timezone.localtime(timezone.now())
    username = account.user.username
    excel_name = '出退勤表　'+str(now.month)+'月.xlsx'
    sheet_path = wageTime_dir/excel_name
    wb = load_workbook(sheet_path)
    try:
        ws = wb[username]
    except Exception as e:
        print("error:")
        print(e)
        return

    user_number = 0
    for sheet in wb.sheetnames:
        if(sheet == username):
            break
        user_number = user_number+1

    print(user_number)

    t_start = account.start_time
    t_end = account.end_time

    account.end_overtime = str(t_end.month)+"/"+str(t_start.day)+" "
    hour = t_end.hour
    minute = 0
    if(t_start.day != t_end.day):
        hour = hour+24
    if(t_end.minute >= 16 and t_end.minute <= 45):
        minute = 30
    elif(t_end.minute >= 46):
        hour = hour+1
    if(hour < 10):
        account.end_overtime += "0"
    account.end_overtime += str(hour)+":"+str(minute)
    if(minute == 0):
        account.end_overtime += "0"
    if(int(((t_end-t_start).total_seconds())/60) < 30):
        account.end_overtime = account.start_overtime
    ws.cell(row=t_start.day+2, column=5, value=account.end_overtime.split()[1])

    send = ""
    if account.is_sending:
        send = "✔"
    ws.cell(row=TodayBehind12(t_start).day+2, column=12, value=send)
    
    print("leaving_sheet_print:")
    print(account.end_overtime)
    # start_str = t_start.strftime("%H:%M")
    # end_str = ""

    # if(t_start.day == t_end.day):
    #     end_str = t_end.strftime("%H:%M")
    # else:
    #     end_str = str(t_end.hour+24)+":"+
    
    account.save()
    wb.save(sheet_path)
    wb.close()

    if(user_number < Account.objects.all().count()):
        
        wb_daily = load_workbook(dailyReport_sheet)
        ws_daily = wb_daily["Revised"]
        
        maching_flg = False
        for title in wb.sheetnames:
            if title == username:
                maching_flg = True
                break

        if maching_flg:
            ws = wb[username]
        else:
            ws = wb.copy_worksheet(wb["ひな形"])
            ws.freeze_panes = 'A3'
            ws.title = username
            ws["E1"] = username
            wb.save(sheet_path)

        UpdateDailyStaff(account.pk)

        wb_daily.close()

def ChangeSheetName(prev_name, name):
    now = timezone.localtime(timezone.now())
    excel_name = '出退勤表　'+str(now.month)+'月.xlsx'
    sheet_path = wageTime_dir/excel_name
    wb = load_workbook(sheet_path)
    ws = wb[prev_name]
    ws.title = name
    wb.save(sheet_path)
    wb.close()

def test(month):
    now = timezone.localtime(timezone.now())

    origin_excel = 'template.xlsx'
    origin_path = wageTime_dir/origin_excel
    origin_wb = xw.Book(origin_path)

    excel_name = '出退勤表　'+str(month)+'月.xlsx'
    sheet_path = wageTime_dir/excel_name
    wb = xw.Book(sheet_path)
    
    origin_wb.sheets["ひな形"].copy(after=wb.sheets[0])

    wb.sheets["Sheet1"].delete()

    origin_wb.save()
    weekDay = ["月","火","水","木","金","土","日"]
    print(calendar.monthrange(2023, month)[1])
    for i in range(calendar.monthrange(2023, month)[1]):
        dates = str(month)+"月"+str(i+1)+"日"
        weekd = weekDay[(date(2023, month, i+1).weekday())]
        wb.sheets["ひな形"].range(3+i,2).value = dates
        wb.sheets["ひな形"].range(3+i,3).value = weekd
        print(dates+weekd)

    app = xw.apps.active

    wb.save()
    origin_wb.close()
    wb.close()

    app.kill()

def makenew(month):

    excel_name = '出退勤表　'+str(month)+'月.xlsx'
    sheet_path = wageTime_dir/excel_name
    wb = xw.Book()

    wb.save(sheet_path)
    app = xw.apps.active
    wb.close()

    app.kill()

def TodayBehind12(datetime):
    delta = timedelta(days=1)
    if datetime.hour < 12:
        datetime -= delta
    print(datetime.day)
    return datetime

def AddSheet(user):
    now = timezone.localtime(timezone.now())
    #xw.App(visible=False)
    excel_name = '出退勤表　'+str(TodayBehind12(now).month)+'月.xlsx'
    sheet_path = wageTime_dir/excel_name
    wb = load_workbook(sheet_path)
    
    wb_daily = load_workbook(dailyReport_sheet)
    ws_daily = wb_daily["Revised"]

    count = Account.objects.all().count()-1

    print(count)
    try:
        ws = wb.copy_worksheet(wb["ひな形"])
        ws.freeze_panes = 'D3'
        ws.title = user
        ws["E1"] = user
        wb.save(sheet_path)

    except:
        print('Write excel error')
    
def DeleteSheet(user):
    now = timezone.localtime(timezone.now())
    excel_name = '出退勤表　'+str(TodayBehind12(now).month)+'月.xlsx'
    sheet_path = wageTime_dir/excel_name
    wb = load_workbook(sheet_path)

    maching_flg = False
    for title in wb.sheetnames:
        if title == user:
            maching_flg = True
            break

    if maching_flg:
        wb.remove(wb[user])
        wb.save(sheet_path)
        
def ConvertOvertimeToDatetime(overdatetime):
    now = timezone.localtime(timezone.now())
    try:
        date = overdatetime.split()[0]
        time = overdatetime.split()[1]
    except:
        return now
    hour = int(time.split(':')[0])
    minute = int(time.split(':')[1])

    convertedDateTime = datetime.strptime(str(now.year)+"/"+date, '%Y/%m/%d')
    if(hour/24 > 0 and hour > 24):
        convertedDateTime = convertedDateTime + timedelta(days=int(hour/24))
        hour %= 24
        
    convertedDateTime = convertedDateTime.replace(hour=hour, minute=minute)

    print("converted:")
    print(convertedDateTime)
    return convertedDateTime

def ConvertDatetimeToOvertime(datetime):
    month = TodayBehind12(datetime).month
    day = TodayBehind12(datetime).day
    hour = datetime.hour
    minute = datetime.minute
    h_zero = ""
    m_zero = ""
    if datetime.hour < 12:
        hour += 24
    if datetime.minute < 10:
        m_zero = "0"

    return str(month)+"/"+str(day)+" "+str(h_zero)+str(hour)+":"+str(m_zero)+str(minute)


is_update = False


def UpdateAccountDrink(pk,date):    
    staff = get_object_or_404(Account,pk=pk)
    staff.staff_drink = 0
    staff.staff_bottle = 0
    staff.save()
    for drink_realtion in staff.sheetstaffrelation_set.all():
        start = timezone.localtime(drink_realtion.checksheet.start_time)
        if TodayBehind12(start).day == date.day:
            staff.staff_drink += drink_realtion.drink
            staff.staff_bottle += drink_realtion.bottle
            staff.save()
    return staff
            

def UpdateAccountBack(pk,date):
    staff = get_object_or_404(Account,pk=pk)
    staff.back = 0
    staff.earnings = 0
    staff.save()
    for relation in staff.sheetaccountrelation_set.all():
        start = timezone.localtime(relation.checksheet.start_time)
        if TodayBehind12(start).day == date.day:
            staff.back += relation.back
            staff.earnings += relation.earnings
            staff.save()
    return staff

def BackCalc(_type,client_num,time):
    back = 0
    match _type:
        case "B":
            back = 300*client_num*time
        case "M":
            print("M")
            if time == 0:
                time = 1
            back = (500+300*(time-1)*client_num)
        case "BJ":
            print("back",type(client_num))
            back = 300*client_num*time
        case "DM":
            if time == 0:
                time = 1
            back = (500+300*(time-1)*client_num)+3000
    print("attr",_type,"num",client_num,"time",time,"back",back)
    return back

def UpdateDaily(date):
    now = timezone.localtime(timezone.now())

    wb = load_workbook(dailyReport_sheet)
    ws = MakeNewDailySheet(wb,dailyReport_sheet,date)
    ws = UpdateDeilyStaff(wb,dailyReport_sheet,date)
    ws = UpdateDilyCheckSheet(wb,dailyReport_sheet,date)

    wb.save(dailyReport_sheet)
    wb.close()

    print("attendance_staff")
    print(date.strftime("%m/%d"))

def MakeNewDailySheet(wb,path,date):
    ws = wb["Revised"]
    daily_day = datetime.strptime(str(ws['B1'].value), '%Y/%m/%d')
    if daily_day.day != TodayBehind12(date).day:
        wb.remove(wb['Revised'])
        ws = wb.copy_worksheet(wb["OriginRevised"])
        ws.title = "Revised"
        ws['B1'] = TodayBehind12(date).strftime("%Y/%m/%d")
        wb.save(path)
    return ws

def UpdateDeilyStaff(wb,path,date):
    i = 13
    j = 5
    ws = wb["Revised"]
    for staff in Account.objects.all():
        start = timezone.localtime(staff.start_time)
        print(TodayBehind12(start).day, date.day)
        if TodayBehind12(start).day == date.day:
            ws['M'+str(i)] = staff.user.username
            try:
                ws['N'+str(i)] = staff.start_overtime.split()[1]
                if staff.start_time <= staff.end_time:
                    ws['O'+str(i)] = staff.end_overtime.split()[1]
            except:
                print("over")
            ws['P'+str(i)] = staff.staff_drink
            ws['Q'+str(i)] = staff.staff_bottle
            ws['R'+str(i)] = staff.debt
            ws['S'+str(i)] = staff.back
            if staff.is_sending:
                ws['J'+str(j)] = staff.user.username
                j += 1
            i += 1
    wb.save(path)
    return ws

def UpdateDilyCheckSheet(wb,path,date):
    i = 0
    j = 0
    ws = wb["Revised"]
    for sheet in CheckSheet.objects.all():
        end = timezone.localtime(sheet.end_time)
        if TodayBehind12(end).day == date.day:
            row = 13+10*(i//5)
            col = 2+2*(i%5)
            ws.cell(row,col,sheet.client_name)
            ws.cell(row,col+1,sheet.client_num)
            ws.cell(row+2,col,sheet.total_fee)
            ws.cell(row+2,col+1,sheet.how_cash[0])
            j = 0
            for staff_relation in sheet.sheetaccountrelation_set.all():
                ws.cell(row+4+j,col,staff_relation.account.user.username)
                ws.cell(row+4+j,col+1,staff_relation.attr)
                j += 1
            i += 1
    wb.save(path)
    return ws


def UpadateAttendanceSheet(pk, date):
    now = timezone.localtime(timezone.now())
    staff = get_object_or_404(Account,pk=pk)
    username = staff.user.username
    excel_name = '出退勤表　'+str(TodayBehind12(now).month)+'月.xlsx'
    sheet_path = wageTime_dir/excel_name

    wb = load_workbook(sheet_path)
    ws = MakeNewStaffSheet(wb,sheet_path,staff.user.username)

    row = date.day+2

    try:
        ws.cell(row,4,staff.start_overtime.split()[1])
        if staff.start_time <= staff.end_time:
            ws.cell(row,5,staff.end_overtime.split()[1])
    except:
        print("over")
    ws.cell(row,7,staff.earnings)

    ws.cell(row,8,staff.staff_drink)
    ws.cell(row,9,staff.staff_bottle)

    ws.cell(row,11,staff.back)

    if staff.is_sending:
        ws.cell(row,12,"✔")

    ws.cell(row,13,staff.debt)
    earnig_total = 0
    for i in range(calendar.monthrange(2023, TodayBehind12(now).month)[1]):
        if ws['G'+str(3+i)].value is not None:
            earnig_total += ws['G'+str(3+i)].value
    #print("earnig_total",earnig_total)

    if earnig_total < 200000:
        wage_index = 4
    elif earnig_total >= 2000000:
        wage_index = 18
    else:
        wage_index = earnig_total//100000+3

    wage_table_path = wageTime_dir/'template.xlsx'
    wage_wb = load_workbook(wage_table_path)
    wage_ws = wage_wb["給料テーブル"]
    wage = wage_ws["C"+str(wage_index)].value
    wage_wb.close()
    print("wage",wage)

    ws.cell(5,16,wage)

    wb.save(sheet_path)
    wb.close()

def MakeNewStaffSheet(wb,path,name):
    maching_flg = False
    for title in wb.sheetnames:
        if title == name:
            maching_flg = True
            break

    if maching_flg:
        ws = wb[name]
    else:
        ws = wb.copy_worksheet(wb["ひな形"])
        ws.freeze_panes = 'A3'
        ws.title = name
        ws["E1"] = name
        wb.save(path)

    return ws



# def ExcelToPDF():
#     excel_name = '出退勤表　'+str(timezone.now().month)+'月.xlsx'
#     sheet_path = wageTime_dir/excel_name
    
#     workbook = load_workbook(sheet_path)
#     worksheet = workbook["tt"]

#     # PDFファイルを作成
#     pdf = FPDF()
#     pdf.add_page()

#     pdf.add_font('IPAGothic', '', '/path/to/IPAGothic.ttf', uni=True)
#     pdf.set_font('IPAGothic', '', 14)


#     # Excelファイルの各セルから値を取得してPDFに書き込む
#     for row in worksheet.iter_rows():
#         for cell in row:
#             pdf.cell(40, 10, str(cell.value))

#     # PDFファイルを保存
#     pdf.output(wageTime_dir/"example1.pdf")
#     #pdf.from_file('test.html', 'pdf1.pdf')
    