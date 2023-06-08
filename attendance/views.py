from django.shortcuts import get_object_or_404, render
from django.views.generic import TemplateView, ListView, FormView, DetailView
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
#from django.contrib.auth.mixins import UserPassesTestMixin
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

import mimetypes
from mysite2.settings import BASE_DIR
from pathlib import Path
#import pythoncom
from . import system
import os
from .forms import AccountForm
from datetime import datetime, timedelta
from .models import Account, Seat, CheckSheet, Item, ItemMenu, SheetAccountRelation, SheetStaffRelation

def Login(request):

    if request.method == 'POST':
        ID = request.POST.get('username')
        Pass = request.POST.get('password')

        user = authenticate(username=ID, password=Pass)
        if user:
            if user.is_superuser:
                login(request, user)
                return HttpResponseRedirect(reverse('AdminForm'))
            
            elif user.is_active:
                login(request, user)
                name = get_object_or_404(Account, user=User.objects.get(username=ID))
                return HttpResponseRedirect(reverse('Result'))
            else:
                return HttpResponse("Not Valid")
        else:
            msg = {
                "incorrect" :True,
            }
            return render(request, 'attendance/login_test.html', msg)
    else:
        return render(request, 'attendance/login_test.html')

@login_required
def Logout(request):
    return HttpResponse("")

@login_required
def Result(request):
    account = get_object_or_404(Account, user=request.user)

    now = timezone.localtime(timezone.now())
    now_12b = system.TodayBehind12(now)
    working_time = now-now
    
    account = system.UpdateAccountDrink(account.pk,now_12b)
    account = system.UpdateAccountBack(account.pk,now_12b)
    account.is_working = not account.is_working
    
    if account.is_working:
        account.start_time = system.AttendanceTimeCalc(now)
        account.start_overtime = system.ConvertDatetimeToOvertime(account.start_time)
        account.end_time = now
        account.end_overtime = ""
        account.is_sending = False
        account.debt = 0
        #system.WriteAttendance(account)
    else:
        if now > account.start_time:
            account.end_time = system.LeavingTimeCalc(now)
            account.end_overtime = system.ConvertDatetimeToOvertime(account.end_time)
            working_time = account.end_time - account.start_time
        else:
            account.end_time = account.start_time
            account.end_overtime = account.start_overtime
        #system.WriteLeaving(account)
    
    account.save()
    logout(request)

    system.UpadateAttendanceSheet(account.pk,now_12b)
    system.UpdateDaily(now_12b)

    params = {#accountでまとめれるね
        "user"          :account.user,
        "is_working"    :account.is_working,
        "time_start"    :account.start_overtime,
        "time_end"      :account.end_overtime,
        "delta_h"       :int(working_time.total_seconds()/3600),
        "delta_m"       :int((working_time.total_seconds()%3600)/60),
    }
    return render(request, "attendance/attendance_form.html", params)

class AdminFrom(ListView):
    model = Account
    template_name = "attendance/admin_form.html"
    def get(self, request):
        return super().get(request)

    def post(self, request):
        print("list:")
        for chk_pk in request.POST.getlist('chk'):
            account = get_object_or_404(Account,pk=chk_pk)
            account.delete()
            User.objects.get(username=account.user.username).delete()
            system.DeleteSheet(account.user.username)
            print(account.user)
        return super().get(request)

class Registration(TemplateView):
    
    def __init__(self):
        self.params = {
        "AccountCreate":False,
        "account_form": AccountForm(),
        }

    def get(self,request):
        if not request.user.is_superuser:
            return HttpResponseRedirect(reverse('Login'))
        else:
            self.params["account_form"] = AccountForm()
            self.params["AccountCreate"] = False
            return render(request,"attendance/register.html",context=self.params)

    def post(self,request):
        if not request.user.is_superuser:
            return HttpResponseRedirect(reverse('Login'))
        else:
            self.params["account_form"] = AccountForm(data=request.POST)
            
            if self.params["account_form"].is_valid():
                account = self.params["account_form"].save()

                account.set_password(account.password)
                account.save()

                now = timezone.now()
                Account.objects.create(
                    user=account,
                    is_working=False,
                    start_time=now,
                    end_time=now,
                    start_overtime="0",
                    end_overtime="0",
                    is_sending=False,
                    staff_drink=0,
                    staff_bottle=0,
                    debt=0,
                    back=0,
                    )

                system.AddSheet(account.username)

                self.params["AccountCreate"] = True

            else:
                print(self.params["account_form"].errors)

            return render(request,"attendance/register.html" ,context=self.params)
        
def DownloadExcel(request,pk):
    now = timezone.localtime(timezone.now())

    account = get_object_or_404(Account,pk=pk)
    user = account.user.username

    path ='excel_sheets/wageTime_sheets 2023/'
    excel_name = '出退勤表　'+str(now.month)+'月.xlsx'
    excel_path = BASE_DIR/path/excel_name
    temp_path = BASE_DIR/path/"download_temp.xlsx"

    wb = load_workbook(excel_path)
    ws = wb[user]

    temp_book = load_workbook(temp_path)
    temp_sheet = temp_book.worksheets[0]
    temp_sheet.title = user
    temp_sheet.freeze_panes = 'D3'
    
    for column in ws.columns:
        temp_sheet.column_dimensions[column[0].column_letter].width = ws.column_dimensions[column[0].column_letter].width
        for cell in column:
            print(cell.column)
            print(cell.number_format)
            temp_cell = temp_sheet.cell(row=cell.row, column=cell.column)
            #temp_sheet[cell.coordinate].border = ws[cell.coordinate].border
            temp_cell.value = cell.value
            _top = Side(style=cell.border.top.style)
            _bottom = Side(style=cell.border.bottom.style)
            _right = Side(style=cell.border.right.style)
            _left = Side(style=cell.border.left.style)
            temp_cell.border = Border(top=_top, bottom=_bottom, right=_right, left=_left)
            temp_cell.alignment = Alignment(horizontal=ws[cell.coordinate].alignment.horizontal)
            
            if temp_cell.column == 2:
                temp_cell.number_format = 'm"月"d"日"'
            else:
                temp_cell.number_format = cell.number_format 
    
    temp_book.save(temp_path)

    #response = HttpResponse(open(path_excel, 'rb').read(), content_type='application/vnd.ms-excel')
    pdf_name = "Time Sheet:"+user+".xlsx"
    print(pdf_name)
    response = HttpResponse(open(temp_path, 'rb').read(), content_type=mimetypes.guess_type(pdf_name)[0])
    
    response['Content-Disposition'] = f'attachment; filename={pdf_name}'
    return response

def AccountEditer(request, pk):
    now = timezone.localtime(timezone.now())
    if request.method == 'POST':
        account = get_object_or_404(Account,pk=pk)
        prev_name = account.user.username
        user = User.objects.get(username=prev_name)

        account.user.username = request.POST.get('username')
        user.username = request.POST.get('username')
        password = request.POST.get('password')
        if request.POST.get('start_t') != None:
            account.start_overtime = request.POST.get('start_t')
            account.start_time = system.ConvertOvertimeToDatetime(account.start_overtime)
        if request.POST.get('end_t') != None:            
            account.start_overtime = request.POST.get('end_t')
            account.end_time = system.ConvertOvertimeToDatetime(account.start_overtime)
        account.debt = request.POST.get('debt')
        account.is_sending = request.POST.get('is_send') == "on"
        
        if prev_name != account.user.username:
            system.ChangeSheetName(prev_name, account.user.username)
        user.save()
        account.save()

        system.UpadateAttendanceSheet(account.pk,system.TodayBehind12(account.start_time))
        system.UpdateDaily(system.TodayBehind12(now))

        print(request.POST.get('start_t'))
        return HttpResponseRedirect(reverse("AdminForm"))
        
    else:
        account = get_object_or_404(Account,pk=pk)
        return render(request,"attendance/account_edit.html", {'ChangedUser':account})
    

def daily(request):
    
    path = BASE_DIR/'excel_sheets/QB Daily Report.xlsx/'
    #path_excel = str(Path(BASE_DIR+dir+'time sheet 4.xlsx').resolve())
    #path_pdf = str(Path(BASE_DIR+dir+'pdf_temp.pdf').resolve())
    
    '''pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")
    _wb = excel.Workbooks.Open(path_excel)
    ws = _wb.Worksheets(user)
    ws.ExportAsFixedFormat(0,path_pdf)
    _wb.Close()
    excel.Quit()

    pythoncom.CoUninitialize() '''
    '''
    wb = load_workbook(dir+'time sheet 4.xlsx')
    ws = wb['s']
    wb.save(dir+'pdf_temp.pdf',SaveFormat.PDF)'''
            
    #response = HttpResponse(open(path_excel, 'rb').read(), content_type='application/vnd.ms-excel')
    pdf_name = "QB Daily Report.xlsx"
    response = HttpResponse(open(path, 'rb').read(), content_type=mimetypes.guess_type(pdf_name)[0])
    
    response['Content-Disposition'] = f'attachment; filename={pdf_name}'
    return response

class SelectSeat(ListView):
    model = Seat
    template_name = "attendance/select_seat.html"
    
    def get(self, request):
        context = {
            "object_list": Seat.objects.all(),
            "CheckSheet_list": CheckSheet.objects.all(),
            }
        return render(request,"attendance/select_seat.html", context)

    def post(self, request):
        now = timezone.localtime(timezone.now())

        month = now.month
        asign_seats = request.POST.getlist('seat-p')
        check_seats = request.POST.getlist('seat-d')

        if "dager_button" in request.POST:
            pk = request.POST.get('dager_button')
            print(pk)
            seat = Seat.objects.get(pk=pk)
            sheet = seat.CheckSheet
            print("red select")

            return HttpResponseRedirect(reverse("CheckSheet", kwargs={'pk':sheet.pk}))

        elif len(asign_seats) != 0:
            seat_list = Seat.objects.filter(pk__in=asign_seats)
            print(seat_list)
            sheet = CheckSheet.objects.create(
                total_fee       =   0,
                discount        =   0,
                how_cash        =   "現金",
                asign           =   False,
                client_name     =   "",
                client_num      =   len(asign_seats),
                start_time      =   now,
                end_time        =   now,
                start_overtime  =   system.ConvertDatetimeToOvertime(now),
                end_overtime    =   "",
                memo_str        =   "",
            )
            staff0 = Account.objects.all()[0]
            SheetAccountRelation.objects.create(checksheet=sheet,account=staff0,attr="B")
            for seat in seat_list:
                seat.CheckSheet = sheet
                seat.is_use = True
                seat.save()
                print(seat)
            manual = get_object_or_404(ItemMenu, menu="Manual")
            for i in range(3):
                item_obj = Item.objects.create(
                    item_name = "",
                    staff = "--",
                    item_num = 0,
                    item_cost = 0,
                    checkSheet = sheet,
                    Menu = manual,
                )
            print(type(sheet.pk))
            return HttpResponseRedirect(reverse("CheckSheet", kwargs={'pk':sheet.pk}))

        else:
            return self.get(request)


        return super().get(request)

class CheckEditer(TemplateView):

    def get(self, request, pk):
        sheet = get_object_or_404(CheckSheet, pk=pk)
        context = {
            "Staff": Account.objects.all(),
            "Relations": sheet.sheetaccountrelation_set.all(),
            "CheckSheet": get_object_or_404(CheckSheet, pk=pk),
            "Menu": get_object_or_404(ItemMenu, menu="Default"),
            }
        return render(request,"attendance/checksheet.html", context)

    def post(self, request, pk):
        now = timezone.localtime(timezone.now())

        item_name_list = request.POST.getlist('item_name')
        drink_list = request.POST.getlist('staff_name')
        item_num_list = request.POST.getlist('item_num')
        item_cost_list = request.POST.getlist('item_cost')
        staff_list = request.POST.getlist('selected_staff')
        staff_attr = request.POST.getlist('staff_attr')
        item_num = len(item_name_list)
        print("staff")
        print(staff_list)
        try:
            check_sheet_obj = get_object_or_404(CheckSheet, pk=pk)
        except:
            return HttpResponseRedirect(reverse("SelectSeat"))
            
        if "cancel" in request.POST:
            print("cancel")
            god = get_object_or_404(CheckSheet, client_name="clientGOD")
            for seat in check_sheet_obj.seat_set.all():
                seat.CheckSheet = god
                seat.is_use = False
                seat.save()
            print(check_sheet_obj)
            check_sheet_obj.delete()
            return HttpResponseRedirect(reverse("SelectSeat"))
        
        print(request.POST.get('total_pay'))
        i = 0

        for _staff in staff_list:
            user = get_object_or_404(User, username=_staff)
            account = get_object_or_404(Account, user=user)
            
            relation = check_sheet_obj.sheetaccountrelation_set.filter(account=account)
            print(relation)
            if relation.exists():
                sheet_account = SheetAccountRelation.objects.get(account=account,checksheet=check_sheet_obj)             
                sheet_account.attr = staff_attr[i]
                time = check_sheet_obj.end_time-check_sheet_obj.start_time
                sheet_account.back = system.BackCalc(sheet_account.attr,check_sheet_obj.client_num,int(time.total_seconds()/3600))
                sheet_account.save()
                print("sheet_account.back")
                print(sheet_account.back)
            else:
                SheetAccountRelation.objects.create(checksheet=check_sheet_obj,account=account,attr=staff_attr[i],back=0)
            
            account = system.UpdateAccountBack(account.pk,system.TodayBehind12(now))
            system.UpadateAttendanceSheet(account.pk,system.TodayBehind12(now))
            i += 1
        i = 0


        print(check_sheet_obj.staff.all())

        for selected_staff in check_sheet_obj.staff.all():
            for _staff in staff_list:
                if selected_staff.user.username == _staff:
                    i = 1
                    break
            if i == 0:
                print("remove")
                user = get_object_or_404(User, username=_staff)
                account = get_object_or_404(Account, user=user)
                relation = SheetAccountRelation.objects.get(account=account,checksheet=check_sheet_obj)
                relation.delete()
            i = 0
            
        print(check_sheet_obj.sheetaccountrelation_set.all().values_list('account', 'checksheet', 'attr'))

        check_sheet_obj.total_fee = request.POST.get('total-f')
        check_sheet_obj.discount = request.POST.get('discount')
        check_sheet_obj.start_overtime = request.POST.get('start_time')
        check_sheet_obj.end_overtime = request.POST.get('end_time')
        check_sheet_obj.start_time = system.ConvertOvertimeToDatetime(check_sheet_obj.start_overtime)
        if check_sheet_obj.end_overtime != "":
            check_sheet_obj.end_time = system.ConvertOvertimeToDatetime(check_sheet_obj.end_overtime)
        if request.POST.get('how_cash') == "現金":
            check_sheet_obj.how_cash = "現金"
        else:
            check_sheet_obj.how_cash = "カード"
        check_sheet_obj.client_name = request.POST.get('client_name')
        check_sheet_obj.client_num = request.POST.get('client_num')
        check_sheet_obj.memo_str = request.POST.get('memo')

        if check_sheet_obj.end_overtime == "":
            check_sheet_obj.asign = True
        check_sheet_obj.save()
        
        i = 0

        for item_obj in check_sheet_obj.item_set.all():
            if item_num > i:
                item_obj.item_name = item_name_list[i]
                item_obj.staff = drink_list[i]
                item_obj.item_num = item_num_list[i]
                item_obj.item_cost = item_cost_list[i]
                print(item_obj)
                item_obj.save()
            else:
                item_obj.delete()
            i += 1

        if i < len(item_name_list):
            for j in range(i,len(item_name_list)):
                new_item = Item.objects.create(
                    item_name = item_name_list[j],
                    staff = drink_list[i],
                    item_num = item_num_list[j],
                    item_cost = item_cost_list[j],
                    checkSheet = check_sheet_obj,
                    Menu = get_object_or_404(ItemMenu, menu="Manual"),
                )
                print(j)
                
        for item_obj in check_sheet_obj.item_set.all():
            if item_obj.staff != "--":
                user = get_object_or_404(User, username=item_obj.staff)
                account = get_object_or_404(Account, user=user)
                relation = check_sheet_obj.sheetstaffrelation_set.filter(account=account)
                print(relation)
                if relation.exists():
                    sheet_account = SheetStaffRelation.objects.get(account=account,checksheet=check_sheet_obj)             
                    
                    if item_obj.item_name == "キャストドリンク":
                        sheet_account.drink = int(item_obj.item_num)
                    else:
                        sheet_account.bottle = int(item_obj.item_cost)*int(item_obj.item_num)
                    sheet_account.save()
                    print("sheet_account.drink")
                    print(sheet_account.drink)
                else:
                    SheetStaffRelation.objects.create(checksheet=check_sheet_obj,account=account,drink=int(item_obj.item_num),bottle=int(item_obj.item_cost))

                account = system.UpdateAccountDrink(account.pk,system.TodayBehind12(now))
                system.UpadateAttendanceSheet(account.pk,system.TodayBehind12(now))

                print("asada",check_sheet_obj.start_time)
                #system.UpadateAttendanceSheet(account.pk,system.TodayBehind12(check_sheet_obj.start_time))
                account.save()

        if "payment" in request.POST:
            if check_sheet_obj.end_overtime == "":
                check_sheet_obj.end_overtime = system.ConvertDatetimeToOvertime(now)
                check_sheet_obj.save()
            return HttpResponseRedirect(reverse("CompSheet", kwargs={'pk':check_sheet_obj.pk}))
        
        if not check_sheet_obj.asign:
            system.UpdateDaily(system.TodayBehind12(now))

        print(check_sheet_obj)
        return HttpResponseRedirect(reverse("SelectSeat"))
  
class CompCheckSheet(TemplateView):
    
    def get(self, request, pk):
        context = {
            "CheckSheet": get_object_or_404(CheckSheet, pk=pk),
            }
        return render(request,"attendance/comp_checksheet.html", context)

    def post(self, request, pk):
        now = timezone.localtime(timezone.now())

        try:
            check_sheet_obj = get_object_or_404(CheckSheet, pk=pk)
        except:
            return HttpResponseRedirect(reverse("SelectSeat"))

        if "cancel" in request.POST:
            check_sheet_obj.end_overtime == ""
            return HttpResponseRedirect(reverse("CheckSheet", kwargs={'pk':check_sheet_obj.pk}))

        god = get_object_or_404(CheckSheet, client_name="clientGOD")
        for seat in check_sheet_obj.seat_set.all():
            seat.CheckSheet = god
            seat.is_use = False
            seat.save()

        print(check_sheet_obj)
        check_sheet_obj.end_time = system.ConvertOvertimeToDatetime(check_sheet_obj.end_overtime)
        check_sheet_obj.asign = False
        check_sheet_obj.save()

        system.UpdateDaily(system.TodayBehind12(now))

        return HttpResponseRedirect(reverse("SelectSeat"))

def DailyEditer(request):
    now = timezone.localtime(timezone.now())
    if request.method == 'POST':
        if "cancel" in request.POST:            
            return HttpResponseRedirect(reverse("SelectSeat"))

        dailyReport_sheet = BASE_DIR/'excel_sheets/QB Daily Report.xlsx'
        wb_daily = load_workbook(dailyReport_sheet)
        ws_daily = wb_daily["Revised"]
        print(type(request.POST.get('envelop')))
        
        ws_daily["D4"] = int(request.POST.get('envelop'))
        ws_daily["F7"] = int(request.POST.get('buyout'))
        ws_daily["H7"] = int(request.POST.get('receipt'))
        ws_daily["B9"] = int(request.POST.get('collect'))
        ws_daily["D9"] = int(request.POST.get('bodyin'))
        ws_daily["F9"] = int(request.POST.get('fare'))
        ws_daily["H9"] = int(request.POST.get('excess'))
        
        wb_daily.save(dailyReport_sheet)
        wb_daily.close()

        #print(request.POST.get('start_t'))
        return HttpResponseRedirect(reverse("SelectSeat"))
        
    else:
        dailyReport_sheet = BASE_DIR/'excel_sheets/QB Daily Report.xlsx'
        wb_daily = load_workbook(dailyReport_sheet, data_only=True)
        ws_daily = wb_daily["Revised"]
        
        total = 0
        for sheet in CheckSheet.objects.all():
            end = timezone.localtime(sheet.end_time)
            if system.TodayBehind12(end).day == system.TodayBehind12(now).day:
                total += sheet.total_fee
        context = {
            "total" : total,
            "envelop" : ws_daily["D4"].value,
            "buyout" : ws_daily["F7"].value,
            "receipt" : ws_daily["H7"].value,
            "collect" : ws_daily["B9"].value,
            "bodyin" : ws_daily["D9"].value,
            "fare" : ws_daily["F9"].value,
            "excess" : ws_daily["H9"].value,
        }
        wb_daily.close()
        return render(request,"attendance/daily_edit.html",context)
    
    

def control(request):
    now = timezone.localtime(timezone.now())
    date = system.UpdateDaily(now)
    print(date)
    #system.UpdateDaily(0)
    
    return render(request,"attendance/outxlsx.html")